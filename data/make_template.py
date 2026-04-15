from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", start_color="FFEB3B")
EXAMPLE_FILL = PatternFill("solid", start_color="E8F5E9")
TEMPLATE_FILL = PatternFill("solid", start_color="FFFFFF")
NOTE_FILL = PatternFill("solid", start_color="FFF9C4")
HEADER_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=10, color="616161")
THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def setup_sheet(ws, headers, widths, note):
    ws["A1"] = note
    ws["A1"].font = NOTE_FONT
    ws["A1"].fill = NOTE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 30
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    for col, w in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def write_row(ws, row_idx, values, fill):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=i, value=v)
        c.fill = fill
        c.alignment = WRAP
        c.border = BORDER


# ============ Sheet 1: shops ============
ws = wb.create_sheet("shops")
setup_sheet(
    ws,
    ["slug", "name", "address", "nearest_station", "business_hours",
     "closed_days", "seat_count", "description", "phone", "is_published"],
    [10, 24, 32, 50, 30, 14, 10, 40, 14, 12],
    "【shopsシート】基本情報。緑の行(mita)は確定済みなので編集不要。"
    "白い行(meguro/sengawa)を埋めてください。nearest_stationは改行可(Alt+Enter)。"
    "is_publishedはTRUE/FALSE。seat_countは数値。"
)

mita_shop = [
    "mita", "ラーメン二郎 三田本店", "東京都港区三田2-16-4",
    "都営地下鉄 浅草線・三田線「三田駅」A3出口より、徒歩約8分\nJR山手線・京浜東北線「田町駅」三田口（西口）より、徒歩約10分",
    "8:30〜20:00", "日曜・祝日", 13,
    "全ての二郎の原点。総帥・山田拓美氏が営む本店。慶應義塾大学のすぐ近く。",
    "", True,
]
meguro_shop = [
    "meguro", "ラーメン二郎 目黒店", "東京都目黒区目黒3-7-2",
    "JR山手線・東京メトロ南北線・都営三田線・東急目黒線「目黒駅」西口より、徒歩約12分\n東急バス「大鳥神社前」バス停より、徒歩約1分",
    "11:00〜16:00 / 18:00〜23:00", "不定休（X(旧Twitter)で告知）", 8,
    "1995年開店の直系二郎老舗。あっさり系の非乳化スープが特徴で、小ラーメン600円と二郎最安クラス。二郎初心者にもおすすめ。",
    "", True,
]
sengawa_shop = [
    "sengawa", "ラーメン二郎 仙川店", "東京都調布市仙川町1-12-2",
    "京王線「仙川駅」より、徒歩約2分（線路沿い）",
    "17:00〜21:00（昼営業なし）", "日曜・祝日", 10,
    "1995年開店の直系二郎。「仙川ブラック」と呼ばれる漆黒の非乳化スープが特徴。醤油と生姜の効いた淡麗系で、細め硬めの平打ちストレート中太麺。",
    "", True,
]
write_row(ws, 3, mita_shop, EXAMPLE_FILL)
write_row(ws, 4, meguro_shop, TEMPLATE_FILL)
write_row(ws, 5, sengawa_shop, TEMPLATE_FILL)
for r in (3, 4, 5):
    ws.row_dimensions[r].height = 80

# ============ Sheet 2: menu_items ============
ws = wb.create_sheet("menu_items")
setup_sheet(
    ws,
    ["shop_slug", "name", "size", "price", "sort_order"],
    [12, 30, 10, 12, 12, 12, 12, 12, 12, 12],
    "【menu_itemsシート】メニュー。緑の行(mita)は確定済み。"
    "meguro/sengawaの行を埋めてください。priceは数値(円は不要)。"
    "sort_orderは1から連番。品数が足りなければ行を追加、余れば削除してOK。"
)

mita_menu = [
    ("mita", "小ラーメン", "小", 700, 1),
    ("mita", "ぶた入り小ラーメン", "小", 850, 2),
    ("mita", "小ぶたダブルラーメン", "小", 950, 3),
    ("mita", "大ラーメン", "大", 750, 4),
    ("mita", "ぶた入り大ラーメン", "大", 900, 5),
    ("mita", "大ぶたダブルラーメン", "大", 1000, 6),
]
row = 3
for m in mita_menu:
    write_row(ws, row, list(m), EXAMPLE_FILL)
    row += 1
meguro_menu = [
    ("meguro", "小ラーメン", "小", 600, 1),
    ("meguro", "大ラーメン", "大", 700, 2),
]
for m in meguro_menu:
    write_row(ws, row, list(m), TEMPLATE_FILL)
    row += 1
sengawa_menu = [
    ("sengawa", "小ラーメン", "小", 800, 1),
    ("sengawa", "小ラーメン豚入り", "小", 900, 2),
    ("sengawa", "大ラーメン", "大", 900, 3),
    ("sengawa", "大ラーメン豚入り", "大", 1000, 4),
]
for m in sengawa_menu:
    write_row(ws, row, list(m), TEMPLATE_FILL)
    row += 1

# ============ Sheet 3: toppings ============
ws = wb.create_sheet("toppings")
setup_sheet(
    ws,
    ["shop_slug", "name", "options", "default_option", "sort_order", "description"],
    [12, 12, 28, 14, 10, 30, 12, 12, 12, 12],
    "【toppingsシート】コール選択肢。緑の行(mita)は確定済み。"
    "optionsはカンマ区切り(例: なし,少なめ,普通,マシ)。"
    "default_optionはoptionsの中の1つ。トッピング種類を増減する場合は行を追加/削除。"
)

default_options = "なし,少なめ,普通,マシ"
mita_top = [
    ("mita", "ニンニク", default_options, "なし", 1, ""),
    ("mita", "ヤサイ", default_options, "普通", 2, ""),
    ("mita", "アブラ", default_options, "普通", 3, ""),
    ("mita", "カラメ", default_options, "普通", 4, ""),
]
row = 3
for t in mita_top:
    write_row(ws, row, list(t), EXAMPLE_FILL)
    row += 1
other_top = [
    ("meguro", "ニンニク", default_options, "なし", 1, ""),
    ("meguro", "ヤサイ", default_options, "普通", 2, ""),
    ("meguro", "アブラ", default_options, "普通", 3, ""),
    ("meguro", "カラメ", default_options, "普通", 4, ""),
    ("sengawa", "ニンニク", default_options, "なし", 1, ""),
    ("sengawa", "ヤサイ", default_options, "普通", 2, ""),
    ("sengawa", "アブラ", default_options, "普通", 3, ""),
    ("sengawa", "カラメ", default_options, "普通", 4, ""),
]
for t in other_top:
    write_row(ws, row, list(t), TEMPLATE_FILL)
    row += 1

# ============ Sheet 4: visit_steps ============
ws = wb.create_sheet("visit_steps")
setup_sheet(
    ws,
    ["shop_slug", "step_number", "title", "description", "tips"],
    [12, 12, 20, 55, 45, 12, 12, 12, 12, 12],
    "【visit_stepsシート】来店フロー7ステップ。緑の行(mita)は確定済み。"
    "meguro/sengawa用に雛形を用意済み。descriptionとtipsを店舗ごとに調整してください。"
    "tipsは空欄可。title変更時はイラスト表示条件(FlowIcons.tsx)に注意。"
)

mita_flow = [
    ("mita", 1, "列に並ぶ",
     "店の外に列ができている場合は最後尾に並ぶ。歩道をふさがないように壁側に1列で。",
     "土曜のランチ時は1時間以上待つことも。平日の夕方が比較的空いている。"),
    ("mita", 2, "食券を購入",
     "並んでいて、看板の下・自販機の前に差し掛かったあたりまで進んだら、食券を買うタイミングです。列を離れ、店内にある券売機で券を購入します。",
     "前の人が買い終わったタイミングを見て券を買いに行きましょう。千円札と小銭のみ対応。事前に崩しておくと安心。"),
    ("mita", 3, "席に着く",
     "店員さんの指示があるまで待つ。指示されたら空いている席に座る。", ""),
    ("mita", 4, "食券をカウンターに置く", "カウンターの上に食券を置く。", ""),
    ("mita", 5, "コールに答える",
     "麺の茹で上がり直前に、店員さんから「○○の方、ニンニク入れますか？」と声がかかります。注文したサイズ（「小の方」「大の方」など）や、座っている位置・目線で呼ばれるので、自分のことだと思ったらトッピングをコールしましょう。",
     "スマホをいじっていたり下を向いていると、声をかけられても気づかないことがあります。茹で上がりが近づいたらスマホはしまって、店員さんの方を見ておきましょう。何も追加しない場合は「そのままで」と伝えればOKです。"),
    ("mita", 6, "食べる",
     "着丼したらなるべく早く食べる。周りのペースに合わせる。",
     "スマホは控えめに。食べることに集中しよう。"),
    ("mita", 7, "丼を上げて退店",
     "食べ終わったら丼をカウンターの上に上げ、台拭きでカウンターを拭いて退店。",
     "テーブルを綺麗にするのがマナー。"),
]
row = 3
for f in mita_flow:
    write_row(ws, row, list(f), EXAMPLE_FILL)
    ws.row_dimensions[row].height = 60
    row += 1
COMMON_STEP2_DESC = "並んでいて、看板の下・自販機の前に差し掛かったあたりまで進んだら、食券を買うタイミングです。列を離れ、店内にある券売機で券を購入します。"
COMMON_STEP2_TIPS = "前の人が買い終わったタイミングを見て券を買いに行きましょう。千円札と小銭のみ対応。事前に崩しておくと安心。"
COMMON_STEP5_DESC = "麺の茹で上がり直前に、店員さんから「○○の方、ニンニク入れますか？」と声がかかります。注文したサイズ（「小の方」「大の方」など）や、座っている位置・目線で呼ばれるので、自分のことだと思ったらトッピングをコールしましょう。"
COMMON_STEP5_TIPS = "スマホをいじっていたり下を向いていると、声をかけられても気づかないことがあります。茹で上がりが近づいたらスマホはしまって、店員さんの方を見ておきましょう。何も追加しない場合は「そのままで」と伝えればOKです。"

meguro_flow = [
    ("meguro", 1, "列に並ぶ",
     "山手通り沿いの店舗前に1列で並ぶ。近隣に住宅があるので、私語は控えめに静かに並ぶのがマナー。",
     "ランチ時は1時間以上の行列になることも。夜の部開始直後が比較的スムーズ。"),
    ("meguro", 2, "食券を購入", COMMON_STEP2_DESC, COMMON_STEP2_TIPS),
    ("meguro", 3, "席に着く", "店員さんの指示があるまで待ち、指示されたらカウンター席に座る。カウンター8席のみで、4人ずつのロット制。", ""),
    ("meguro", 4, "食券をカウンターに置く", "カウンターの上に食券を置いて待つ。", ""),
    ("meguro", 5, "コールに答える", COMMON_STEP5_DESC, COMMON_STEP5_TIPS),
    ("meguro", 6, "食べる", "着丼したらなるべく早く食べる。ロット制なので同じ組の人とペースを合わせるとスマート。", ""),
    ("meguro", 7, "丼を上げて退店", "食べ終わったら丼をカウンターの上に上げ、台拭きでカウンターを拭いて退店する。", "次の客のために席をきれいにするのがマナー。"),
]
sengawa_flow = [
    ("sengawa", 1, "列に並ぶ",
     "京王線の線路沿いにある店舗前に1列で並ぶ。駅から徒歩2分。",
     "夜のみの営業（17:00〜21:00）。営業開始前から行列ができることが多い。"),
    ("sengawa", 2, "食券を購入", COMMON_STEP2_DESC, COMMON_STEP2_TIPS),
    ("sengawa", 3, "席に着く", "店員さんの指示に従って空いた席に座る。カウンター10席のみで、4人・6人のロット制。", ""),
    ("sengawa", 4, "食券をカウンターに置く", "着席したらカウンターの上に食券を置く。麺量の調整（麺少なめなど）を希望する場合はここで伝える。", ""),
    ("sengawa", 5, "コールに答える", COMMON_STEP5_DESC, COMMON_STEP5_TIPS),
    ("sengawa", 6, "食べる", "着丼したら速やかに食べる。仙川ブラックと呼ばれる漆黒の非乳化スープを堪能しよう。", ""),
    ("sengawa", 7, "丼を上げて退店", "食べ終わったら丼をカウンターの上に上げ、台拭きでカウンターを拭いて退店。", ""),
]
for f in meguro_flow + sengawa_flow:
    write_row(ws, row, list(f), TEMPLATE_FILL)
    ws.row_dimensions[row].height = 60
    row += 1

out = r"C:\Users\yykkm\Desktop\PJ\Jiro_training\data\jiro-shops-template.xlsx"
wb.save(out)
print(f"saved: {out}")
print(f"sheets: {wb.sheetnames}")
